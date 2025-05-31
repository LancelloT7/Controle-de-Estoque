from django.contrib import admin
from .models import Item, Fabricante, NotaFiscal, HistoricoValorItem
from django.utils.html import format_html
from entrada_de_itens.models import Entrada
from saida_de_itens.models import Saida
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse


# Register your models here.
class HistoricoValorItemInline(admin.TabularInline):
    model = HistoricoValorItem
    extra = 0
    can_delete = False
    readonly_fields = ['valor_anterior', 'valor_novo', 'responsavel', 'data_alteracao']
    verbose_name = "Histórico de Valor"
    verbose_name_plural = "Histórico de Valores" 

class EntradaInline(admin.TabularInline):
    model = Entrada
    extra = 0
    readonly_fields = ['nota_fiscal', 'responsavel', 'data', 'quantidade']
    can_delete = False
    verbose_name_plural = 'Entradas relacionadas'

class SaidaInline(admin.TabularInline):
    model = Saida
    extra = 0
    readonly_fields = ['responsavel', 'data', 'quantidade']
    can_delete = False
    verbose_name_plural = 'Saídas relacionadas'    

@admin.register(Item)
class ItemAdmin(admin.ModelAdmin):
    search_fields = ['fabricante__nome_do_fabricante', 'nome',]
    list_display = ['sku', 'nome', 'modelo', 'partnumber', 'fabricante' ,'quantidade', 'disponivel', 'endereco' ,'preview',  'valor_atual', 'total_estoque_admin', 'ultima_saida']
    list_filter = ['sku', 'disponivel',  'modelo', 'fabricante__nome_do_fabricante', 'nome', 'partnumber', 'ultima_saida']
    inlines = [EntradaInline, SaidaInline, HistoricoValorItemInline]
    actions = ['exportar_para_excel']

    def total_estoque_admin(self, obj):
        return obj.total_estoque()
    total_estoque_admin.short_description = 'Total Estoque'

    def save_model(self, request, obj, form, change):
        if obj.pk:
            item_antigo = Item.objects.get(pk=obj.pk)
            if item_antigo.valor_atual != obj.valor_atual:
                HistoricoValorItem.objects.create(
                    item=obj,
                    valor_anterior=item_antigo.valor_atual,
                    valor_novo=obj.valor_atual,
                    responsavel=request.user
                )
        super().save_model(request, obj, form, change)

    def preview(self, obj):
        if obj.img:
            return format_html(
                '<img src="{}" style="max-height: 100px; max-width: 150px; object-fit: contain;" />',
                obj.img.url
            )
        return "-"
    preview.short_description = 'Imagem'
    def has_change_permission(self, request, obj=None):
        # Permite acesso à página de edição
        return True

    def get_readonly_fields(self, request, obj=None):
        if request.user.groups.filter(name='Funcionarios').exists():
            # Se for funcionário, todos os campos são readonly, exceto "endereco"
            all_fields = [field.name for field in self.model._meta.fields]
            return [f for f in all_fields if f != 'endereco']
        return super().get_readonly_fields(request, obj)


    def exportar_para_excel(self, request, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Estoque"

        # Definir cabeçalhos (incluindo Total Estoque)
        headers = [
            'SKU', 'Nome', 'Modelo', 'PartNumber', 'Fabricante',
            'Quantidade', 'Disponível', 'Endereço', 'Valor Atual', 'Total Estoque'
        ]
        ws.append(headers)

        # Estilos para cabeçalho
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")  # Azul forte
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col_num, column_title in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Preencher dados
        fill_odd = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")  # linha clara
        fill_even = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # linha branca

        for row_num, item in enumerate(queryset, 2):
            fill = fill_odd if row_num % 2 == 0 else fill_even

            # Calcula total estoque (quantidade * valor_atual)
            total_estoque = item.quantidade * (item.valor_atual or 0)

            row = [
                item.sku,
                item.nome,
                item.modelo,
                item.partnumber,
                item.fabricante.nome_do_fabricante if item.fabricante else "",
                item.quantidade,
                "Sim" if item.disponivel else "Não",
                item.endereco,
                float(item.valor_atual) if item.valor_atual is not None else 0,
                float(total_estoque),
            ]

            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.fill = fill
                # Alinhamentos e formatações
                if col_num in [6, 10]:  # quantidade e total estoque alinhados à direita (número)
                    cell.alignment = Alignment(horizontal="right")
                    if col_num == 10:
                        cell.number_format = 'R$ #,##0.00'  # formato moeda para total estoque
                    elif col_num == 6:
                        cell.number_format = '0'  # formato número inteiro para quantidade
                elif col_num == 7:  # disponível centralizado
                    cell.alignment = Alignment(horizontal="center")
                elif col_num == 9:  # valor atual alinhado à direita com formato moeda
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = 'R$ #,##0.00'
                else:
                    cell.alignment = Alignment(horizontal="left")

                cell.border = thin_border

        # Ajustar largura das colunas automaticamente
        for col_num, column_title in enumerate(headers, 1):
            max_length = len(column_title)
            column_letter = get_column_letter(col_num)
            for cell in ws[column_letter]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width

        # Congelar primeira linha
        ws.freeze_panes = "A2"

        # Resposta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=relatorio_estoque.xlsx'
        wb.save(response)
        return response


@admin.register(NotaFiscal)
class NotaFiscalAdmin(admin.ModelAdmin):
    search_fields = ['numero_de_nota']
    list_display = ['numero_de_nota']

@admin.register(Fabricante)
class FabricanteAdmin(admin.ModelAdmin):
    search_fields = ['nome_do_fabricante']
    list_display = ['nome_do_fabricante']



@admin.register(HistoricoValorItem)
class HistoricoValorItemAdmin(admin.ModelAdmin):
    search_fields = ['item__sku', 'valor_anterior', 'valor_novo', 'data_alteracao']
    list_display = ['valor_anterior', 'valor_novo', 'responsavel','data_alteracao']
    list_filter = ['item__sku', 'item__nome', 'item__partnumber', 'data_alteracao']
    
    

